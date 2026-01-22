import streamlit as st
import datetime
import re
import time
import math
import pandas as pd
import requests
from bs4 import BeautifulSoup
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Studico.", layout="wide", page_icon="🎓")

# --- CUSTOM CSS ---
st.markdown("""
    <style>
    /* --- TOMBOL UTAMA (Primary) --- */
    div.stButton > button[kind="primary"]:not(:disabled), 
    div.stDownloadButton > button[kind="primary"]:not(:disabled) {
        background-color: #2D3E50; 
        color: white;
        border: none;
    }
    
    div.stButton > button[kind="primary"]:not(:disabled):hover, 
    div.stDownloadButton > button[kind="primary"]:not(:disabled):hover {
        background-color: #1A2530; 
        color: white;
        border: none;
    }
    
    div.stButton > button[kind="primary"]:not(:disabled):focus, 
    div.stDownloadButton > button[kind="primary"]:not(:disabled):focus {
        box-shadow: none;
        color: white;
    }

    div.stButton > button[kind="primary"]:disabled, 
    div.stDownloadButton > button[kind="primary"]:disabled {
        background-color: rgba(45, 62, 80, 0.4); 
        color: rgba(255, 255, 255, 0.4); 
        border: 1px solid rgba(255, 255, 255, 0.1); 
        cursor: not-allowed;
    }

    /* --- NUMBER INPUT (+/- Buttons) --- */
    div[data-testid="stNumberInput"] button {
        color: #2D3E50 !important;
        border-color: rgba(45, 62, 80, 0.2) !important;
    }
    div[data-testid="stNumberInput"] button:hover {
        border-color: #2D3E50 !important;
        background-color: rgba(45, 62, 80, 0.05) !important;
    }
    div[data-testid="stNumberInput"] button:active,
    div[data-testid="stNumberInput"] button:focus,
    div[data-testid="stNumberInput"] button:focus-visible {
        background-color: #2D3E50 !important;
        color: white !important;
        border-color: #2D3E50 !important;
        box-shadow: none !important; /* Hilangkan glow oren default */
        outline: none !important;
    }

    /* --- TABS CUSTOMIZATION (DEFAULT / LIGHT MODE) --- */
    div[data-baseweb="tab-highlight"] {
        background-color: #2D3E50 !important;
    }
    button[data-baseweb="tab"][aria-selected="true"] {
        color: #2D3E50 !important;
    }
    button[data-baseweb="tab"]:hover {
        color: #2D3E50 !important;
        background-color: transparent !important;
    }
    button[data-baseweb="tab"]:focus {
        outline: none !important;
    }

    /* --- DARK MODE OVERRIDES --- */
    @media (prefers-color-scheme: dark) {
        /* Tabs */
        div[data-baseweb="tab-highlight"] {
            background-color: #E2E8F0 !important; 
        }
        button[data-baseweb="tab"][aria-selected="true"] {
            color: #E2E8F0 !important;
        }
        button[data-baseweb="tab"]:hover {
            color: #F8FAFC !important; 
        }

        /* Number Input (+/- Buttons) - Dark Mode */
        div[data-testid="stNumberInput"] button {
            color: #E2E8F0 !important;
            border-color: rgba(226, 232, 240, 0.2) !important;
        }
        div[data-testid="stNumberInput"] button:hover {
            border-color: #E2E8F0 !important;
            background-color: rgba(226, 232, 240, 0.1) !important;
        }
        div[data-testid="stNumberInput"] button:active,
        div[data-testid="stNumberInput"] button:focus,
        div[data-testid="stNumberInput"] button:focus-visible {
            background-color: #E2E8F0 !important;
            color: #000000 !important;
            border-color: #E2E8F0 !important;
            box-shadow: none !important;
            outline: none !important;
        }
    }
    </style>
""", unsafe_allow_html=True)

# --- Helper Function: Custom Toast ---
def show_custom_toast(message, type="error", duration=5):
    unique_id = int(time.time() * 1000)
    
    if type == "error":
        bg_color = "#FFE9E9" 
        text_color = "#991B1B" 
        icon = "⚠️"
    else:
        bg_color = "#E8F9EE" 
        text_color = "#065F46" 
        icon = "✅"
    
    # CSS Animation for fade out
    html_code = f"""
    <style>
        @keyframes slideInDown-{unique_id} {{
            0% {{ opacity: 0; top: -50px; }}
            10% {{ opacity: 1; top: 90px; }} 
            90% {{ opacity: 1; top: 90px; }}
            100% {{ opacity: 0; top: -50px; pointer-events: none; }}
        }}
        .custom-toast-{unique_id} {{
            position: fixed;
            right: 20px;
            top: 90px;
            background-color: {bg_color};
            color: {text_color};
            padding: 12px 24px;
            border-radius: 8px;
            box-shadow: 0 4px 12px rgba(0,0,0,0.1);
            z-index: 999999;
            font-family: 'Source Sans Pro', sans-serif;
            display: flex;
            align-items: center;
            gap: 12px;
            animation: slideInDown-{unique_id} {duration}s ease-in-out forwards;
        }}
    </style>
    <div class="custom-toast-{unique_id}">
        <span style="font-size: 1.5rem;">{icon}</span>
        <div style="font-weight: 500;">{message}</div>
    </div>
    """
    return html_code

# --- Helper Function: Scraper  ---
def scrape_dicoding_syllabus(url):
    try:
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8",
            "Accept-Language": "en-US,en;q=0.9,id;q=0.8",
            "Referer": "https://www.dicoding.com/",
            "Upgrade-Insecure-Requests": "1"
        }
        # 1. Ambil HTML dari Link
        response = requests.get(url, headers=headers)
        response.raise_for_status()
        
        # 2. Proses HTML 
        soup = BeautifulSoup(response.text, 'html.parser')
        
        # Ambil Nama Kelas 
        class_name_tag = soup.find('h3', class_="mb-3 font-weight-bold")
        if not class_name_tag:
            class_name_tag = soup.find('h1')
            
        class_name = class_name_tag.get_text(strip=True) if class_name_tag else "Dicoding Class"
        
        modules = []
        
        # --- LOGIC USER ---
        # Loop tiap syllabus-category (modul)
        syllabus_categories = soup.select("div.syllabus-category")
        
        if not syllabus_categories:
            return None, "Gagal menemukan elemen silabus. Pastikan link publik dan benar."

        for cat in syllabus_categories:
            modul_title_tag = cat.find("h5", class_="syllabus-category__title")
            modul_title = modul_title_tag.get_text(strip=True) if modul_title_tag else "Tanpa Modul"
            
            current_articles = []
            
            for li in cat.select("li"):
                judul, menit = None, None

                a_tag = li.find("a")
                p_tag = li.find("p", class_="mb-0 text-secondary")

                if a_tag and p_tag:
                    judul = a_tag.get_text(strip=True)
                    menit_text = p_tag.get_text(strip=True)
                    match = re.search(r'(\d+)', menit_text)
                    if match:
                        menit = match.group(1)

                if not judul: 
                    judul_p = li.find("p", class_="syllabus-module-list__link")
                    menit_p = li.find("p", class_="mb-0 text-secondary")

                    if judul_p and menit_p:
                        judul = judul_p.get_text(strip=True)
                        menit_text = menit_p.get_text(strip=True)
                        match = re.search(r'(\d+)', menit_text)
                        if match:
                            menit = match.group(1)

                # Jika berhasil dapet judul dan menit
                if judul and menit:
                    try:
                        duration_val = int(menit)
                        current_articles.append({"title": judul, "duration": duration_val})
                    except ValueError:
                        pass 

            # Simpan modul kalo ada artikelnya
            if current_articles:
                modules.append({"name": modul_title, "articles": current_articles})

        if not modules:
            return None, "Link bisa diakses tapi tidak ditemukan modul/artikel. Cek apakah silabus kosong?"
            
        return {"name": class_name, "modules": modules}, None

    except Exception as e:
        return None, f"Error: {str(e)}"

# --- Step 0: Initialize session state ---
if "classes" not in st.session_state:
    st.session_state.classes = []
if "schedule" not in st.session_state:
    st.session_state.schedule = {}

# ========== SIDEBAR ==========
with st.sidebar:
    st.write("# ⚙️ Config & Input")

    # --- Step 1: Set Date & Time ---
    st.markdown("---")
    st.write("## 1. Set Date & Time")
    
    start_date = st.date_input("Start Date", datetime.date.today())
    end_date = st.date_input("End Date")

    if end_date < start_date:
        st.error("End Date tidak boleh lebih awal dari Start Date.", icon="⚠️")
    
    st.caption("Target Belajar per Hari:")
    col_h, col_m = st.columns(2)
    with col_h:
        target_hours = st.number_input("Hour(s)", min_value=0, max_value=24, value=3, step=1)
    with col_m:
        target_minutes_input = st.number_input("Minute(s)", min_value=0, max_value=59, value=30, step=5)
        
    minutes_per_day = (target_hours * 60) + target_minutes_input
    st.info(f"Target: **{target_hours}h {target_minutes_input}m** / day | **{minutes_per_day}m** / day")
 
    # --- Step 2: Input Methods ---
    st.markdown("---")
    st.write("## 2. Add Materials")
    
    input_method = st.radio("Metode Input:", ["🔗 Auto Scrape (URL)", "✍️ Manual Input"])
    
    # === METHOD A: AUTO SCRAPE ===
    if input_method == "🔗 Auto Scrape (URL)":
        st.markdown("Pastikan link yang dimasukkan adalah halaman **Silabus** atau **Detail Kelas** Dicoding yang publik.")
        url_input = st.text_input("Dicoding Class URL", placeholder="https://www.dicoding.com/academies/...")
        
        if st.button("🚀 Scrape & Add", type="primary", use_container_width=True):
            if url_input:
                with st.spinner("Sedang membaca silabus..."):
                    result, error = scrape_dicoding_syllabus(url_input)
                    
                    if result:
                        # Cek duplikat kelas
                        if any(c['name'] == result['name'] for c in st.session_state.classes):
                             st.warning(f"Kelas '{result['name']}' sudah ada di list.", icon="⚠️")
                        else:
                            st.session_state.classes.append(result)
                            st.toast(f"Berhasil menambahkan: {result['name']}", icon="✅")
                            time.sleep(2)
                            st.rerun()
                    else:
                        st.error(f"Gagal scraping: {error}", icon="❌")
            else:
                st.warning("Masukkan URL dulu yaa.", icon="⚠️")

    # === METHOD B: MANUAL INPUT ===
    else:
        # Input Class
        class_input = st.text_input("Enter Class Name", key="class_input_man", placeholder="e.g., Belajar Python Dasar")
        if st.button("➕ Add Class", key="save_class_man", use_container_width=True):
            if class_input.strip():
                if any(c['name'].lower() == class_input.strip().lower() for c in st.session_state.classes):
                    st.warning(f"Kelas '{class_input.strip()}' sudah ada di list.", icon="⚠️")
                else:
                    st.session_state.classes.append({"name": class_input.strip(), "modules": []})
                    st.toast(f"Berhasil menambahkan: {class_input.strip()}", icon="✅")
                    time.sleep(2)
                    st.rerun()
            else:
                st.warning("Nama kelas tidak boleh kosong.", icon="⚠️")

    # --- LIST CLASSES & EDIT MODULES  ---
    st.write("### 📚 Class List")
    
    if not st.session_state.classes:
        st.caption("*Belum ada kelas. Tambahkan via URL atau Manual.*")

    # Loop classes untuk manage Modules/Articles
    for class_idx, class_item in enumerate(st.session_state.classes):
        with st.expander(f"📘 {class_item['name']}", expanded=False):
            # Tombol Hapus Kelas
            if st.button(f"🗑️ Delete Class", key=f"del_class_{class_idx}"):
                st.session_state.classes.pop(class_idx)
                st.session_state.schedule = {} # Reset schedule saat kelas dihapus
                st.rerun()

            st.markdown("#### Modules")
            
            # Form tambah modul manual
            with st.form(key=f"add_mod_form_{class_idx}"):
                c1, c2 = st.columns([3, 1])
                with c1:
                    new_mod_name = st.text_input("Enter Module Name", placeholder="Modul X")
                with c2:
                    st.text("") # spacer
                    st.text("") 
                    add_mod_btn = st.form_submit_button("➕ Add")
                
                if add_mod_btn and new_mod_name:
                    class_item["modules"].append({"name": new_mod_name, "articles": []})
                    st.rerun()

            # Loop Modules
            for module_idx, module_item in enumerate(class_item["modules"]):
                st.markdown(f"**📂 {module_item['name']}**")
                
                # Input artikel manual (Bulk)
                input_val = st.text_area(
                    f"Edit/Add Articles (Format: Judul [spasi] Menit)",
                    value="",
                    placeholder="Pengenalan 5\nInstalasi Tools 15",
                    key=f"area_{class_idx}_{module_idx}",
                    height=68
                )
                
                col_act1, col_act2 = st.columns(2)
                with col_act1:
                    if st.button("Add to List", key=f"btn_add_art_{class_idx}_{module_idx}"):
                        for line in input_val.splitlines():
                            match = re.match(r"(.+?)\s+(\d+)$", line.strip())
                            if match:
                                t, d = match.groups()
                                module_item["articles"].append({"title": t.strip(), "duration": int(d)})
                        st.rerun()
                
                # List Artikel yang sudah ada
                if module_item["articles"]:
                    for art_idx, art in enumerate(module_item["articles"]):
                        c_text, c_del = st.columns([4, 1])
                        with c_text:
                            st.caption(f"• {art['title']} ({art['duration']}m)")
                        with c_del:
                            if st.button("❌", key=f"del_art_{class_idx}_{module_idx}_{art_idx}"):
                                module_item["articles"].pop(art_idx)
                                st.rerun()
                else:
                    st.caption("*Belum ada artikel*")
                st.divider()

    # --- Step 3: Generate Schedule ---
    st.markdown("---")
    st.write("## 3. Generate")
    
    can_generate = (
        end_date >= start_date and 
        st.session_state.classes and 
        minutes_per_day > 0
    )
    
    if st.button("📅 Generate Schedule", type="primary", disabled=not can_generate, use_container_width=True):
        all_tasks = []
        for class_item in st.session_state.classes:
            for module_item in class_item["modules"]:
                for article in module_item["articles"]:
                    all_tasks.append({
                        "class": class_item["name"],
                        "module": module_item["name"],
                        "title": article["title"],
                        "duration": article["duration"]
                    })

        total_days = (end_date - start_date).days + 1
        schedule = {start_date + datetime.timedelta(days=i): [] for i in range(total_days)}

        current_day = start_date
        used_minutes = 0

        # Greedy Algorithm
        task_idx = 0
        total_tasks = len(all_tasks)
        
        while task_idx < total_tasks and current_day <= end_date:
            task = all_tasks[task_idx]
            
            remaining_space = minutes_per_day - used_minutes
            
            if task["duration"] <= remaining_space:
                schedule[current_day].append(task)
                used_minutes += task["duration"]
                task_idx += 1
            else:
                if used_minutes == 0:
                    schedule[current_day].append(task)
                    current_day += datetime.timedelta(days=1)
                    used_minutes = 0
                    task_idx += 1
                else:
                    current_day += datetime.timedelta(days=1)
                    used_minutes = 0
        
        if task_idx < total_tasks:
            msg = "<b>Waktunya gak cukup nih.</b> <br><span style='font-size: 0.9em; opacity: 0.9;'>Coba perpanjang End Date atau tambah durasi belajar, lalu generate ulang.</span>"
            st.markdown(show_custom_toast(msg, type="error", duration=10), unsafe_allow_html=True)
        else:
            st.markdown(show_custom_toast("Jadwal Berhasil Dibuat!", type="success", duration=5), unsafe_allow_html=True)
            st.balloons()

        st.session_state.schedule = schedule


# ========== MAIN CONTENT ==========
st.title("Studico.")
st.markdown(f"##### Set your Dicoding study schedule automatically.")

tab1, tab2, tab3 = st.tabs(["🗓️ Preview", "📝 Markdown", "📥 Excel"])

# Preview
with tab1:
    if st.session_state.schedule:
        total_items = sum(len(v) for v in st.session_state.schedule.values())
        st.metric("Total Item Dijadwalkan", total_items)
        
        for day, tasks in st.session_state.schedule.items():
            total_minutes = sum(t["duration"] for t in tasks)
            total_hours = total_minutes / 60
            
            if total_minutes == 0:
                card_color = "grey"
                status = "🏖️ Free Day"
            elif total_minutes > minutes_per_day:
                card_color = "red" 
                status = "🔥 Overload"
            else:
                card_color = "green"
                status = "✅ On Track"

            with st.expander(f"{day.strftime('%A, %d %b %Y')} | {status} ({total_minutes} min)", expanded=(day == start_date)):
                if tasks:
                    for t in tasks:
                        st.markdown(f"- **{t['class']}** / *{t['module']}* / {t['title']} `({t['duration']} min)`")
                else:
                    st.write("Istirahat dulu bro..")
    else:
        st.info("👈 Set date & time, add materials, dan generate schedule dulu di sidebar.")

# Markdown
with tab2:
    if st.session_state.schedule:
        st.subheader("Copy-Paste ke Notion")
        st.caption("Copy text di bawah, lalu paste di Notion untuk membuat to-do-list.")
        
        markdown_text = ""
        
        for day, tasks in st.session_state.schedule.items():
            total_minutes = sum(t["duration"] for t in tasks)
            if tasks: 
                date_str = day.strftime('%A, %d %B %Y')
                markdown_text += f"### 📅 {date_str}\n"
                markdown_text += f"**Target:** {total_minutes} min\n\n"
                for t in tasks:
                    markdown_text += f"- [ ] **{t['class']}** | *{t['module']}* | {t['title']} ({t['duration']}m)\n"
                markdown_text += "\n"

        st.code(markdown_text, language="markdown")
    else:
        st.info("👈 Generate schedule dulu bro.")

# Excel
with tab3:
    if st.session_state.schedule:
        # --- Export Excel ---
        export_data = []
        for day, tasks in st.session_state.schedule.items():
            total_minutes_day = sum(t["duration"] for t in tasks) if tasks else 0
            
            if tasks:
                for t in tasks:
                    export_data.append({
                        "Date": day.strftime("%d-%m-%Y"),
                        "Class": t["class"],
                        "Module": t["module"],
                        "Article": t["title"],
                        "Duration (min)": t["duration"],
                        "Total Duration (min/day)": total_minutes_day,
                        "Status (✅)": "☐"
                    })

        df_export = pd.DataFrame(export_data)
        
        if not df_export.empty:
            # --- Save as Excel ---
            output = BytesIO()
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                df_export.to_excel(writer, index=False, sheet_name="Schedule")
            output.seek(0)

            wb = load_workbook(output)
            ws = wb["Schedule"]

            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                 top=Side(style='thin'), bottom=Side(style='thin'))

            def merge_column_in_rows(ws, col_letter, start_row, end_row):
                if end_row > start_row:
                    ws.merge_cells(f"{col_letter}{start_row}:{col_letter}{end_row}")
                    ws[f"{col_letter}{start_row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            def merge_column(ws, col_letter, start_row, end_row):
                row = start_row
                while row <= end_row:
                    merge_start = row
                    while row + 1 <= end_row and ws[f"{col_letter}{row}"].value == ws[f"{col_letter}{row+1}"].value:
                        row += 1
                    merge_column_in_rows(ws, col_letter, merge_start, row)
                    row += 1

            current_row = 2
            while current_row <= ws.max_row:
                date_value = ws[f"A{current_row}"].value
                start_row = current_row
                while current_row + 1 <= ws.max_row and ws[f"A{current_row+1}"].value == date_value:
                    current_row += 1
                end_row = current_row

                # A=Date, B=Class, C=Module, D=Article, E=Duration, F=Total, G=Status
                merge_column_in_rows(ws, "A", start_row, end_row) # Merge Date
                merge_column(ws, "B", start_row, end_row)         # Merge Class
                merge_column(ws, "C", start_row, end_row)         # Merge Module
                merge_column_in_rows(ws, "F", start_row, end_row) # Merge Total Duration

                current_row += 1

            # Apply border + alignment + wrap_text
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.border = thin_border
                    col_letter = get_column_letter(cell.column)
                    if col_letter == "D" and cell.row != 1: 
                        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                    else:
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            output_merged = BytesIO()
            wb.save(output_merged)
            excel_data = output_merged.getvalue()

            # --- button download ---
            col1, col2 = st.columns([5, 2])
            with col1:
                st.subheader("Preview Excel Table")
                st.caption("Download file excel melalui tombol di samping.")
            with col2:
                st.download_button(
                    label="📥 Download Excel",
                    data=excel_data,
                    file_name=f"Studico._{start_date}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary"
                )

            # --- Preview table ---
            st.dataframe(df_export, use_container_width=True)
        else:
            st.warning("Jadwal kosong atau belum digenerate.")

    else:
        st.info("👈 Generate schedule dulu bro.")

# --- FOOTER / WATERMARK ---
st.markdown("---")
st.markdown("""
    <div style="text-align: center; color: #6c757d; padding: 10px;">
        <small>Created by <b>serafiua</b> | Powered by Streamlit</small>
    </div>
""", unsafe_allow_html=True)
